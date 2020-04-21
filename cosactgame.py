import openpyxl
from tabulate import tabulate
import random

# import pyinputplus as pyip

score = 0

wb = openpyxl.load_workbook("Book1.xlsx")
# print([x for x in list(wb)])
ws = wb.worksheets[0]
print([ws.cell(row=1, column=x).value for x in range(1, ws.max_column + 1)])

# use a loop for finding all information into the requried format
headerList = [ws.cell(row=1, column=x).value for x in range(1, ws.max_column + 1)]
# print(headerList)

section = list()
for counter1 in range(2, ws.max_row, 1):
    section.append({})  # making an empty dictionary here
    dict = section[counter1 - 2]  # for simplyfying the below alloting to dict
    for counter in range(1, 6):
        dict[headerList[counter - 1]] = ws.cell(row=counter1, column=counter).value
    section[counter1 - 2] = dict  # alloting to section the dict
    # print(section[counter1-2])

# print(f"\n {section[300]} ")                                  #for printing random section for checking

# for converting the headerList to be supplied to the tabulate command
headerList1 = headerList[0:5]  # for taking the first 5 columns only
headerList2 = {}
for counter2 in range(0, 5):
    headerList2[headerList1[counter2]] = headerList1[counter2]

# this command will take care to print the whole table
# print(tabulate(section, headerList2, tablefmt="psql"))

# ______________________________________________________________________________________________________________________

# todo: randomly select any 10 section numbers and store them in a dictionary
# #todo: randomly select any 10 numbers and store them in a dictionary.this is for the questions to be asked question [0-9]
questionPaper = [{} for _ in range(0, 10)]
selectRandomSections = [random.randrange(1, ws.max_row - 1) for _ in range(0, 10)]
print(selectRandomSections)
for counter in range(0, 10):

    a = section[
        selectRandomSections[counter]
    ]  # assigning to a and q to simplify the below statements
    q = questionPaper[counter]

    q["selectedSerial"] = a.get("serial")
    q["selectedSectionNumber"] = a.get("sectionNumber")
    q["selectedTextOfSection"] = a.get("textOfSection")
    # wrongAnswerSections = []
    # for counter in range(0, 3):
    #     wrongAnswerSections.append(selectRandomSections[counter + random.randrange(1, 40)])
    q["wanswer1"] = section[random.randrange(1, ws.max_row - 1)].get("sectionNumber")
    q["wanswer2"] = section[random.randrange(1, ws.max_row - 1)].get("sectionNumber")
    q["wanswer3"] = section[random.randrange(1, ws.max_row - 1)].get("sectionNumber")
    q["selectedAnswer"] = 0

    questionPaper[counter] = q

print(tabulate(questionPaper))

for counter in range(0, 10):
    # a=(questionPaper[counter].get("selectedTextOfSection"))
    a = questionPaper[counter]

    d = list((a["selectedSectionNumber"], a["wanswer1"], a["wanswer2"], a["wanswer3"]))
    c = sorted(d)
    # print(a)
    b = a["selectedTextOfSection"]
    # print(type(a),type(b),type(c),type(d))
    # print(a,b,c,d)
    e = a["selectedSectionNumber"]
    xyz=input(
        f"{counter+1} To which section of the Companies Act 2013 does the below text pertain: \n    {b} \nA) {c[0]} \t B) {c[1]} \t C) {c[2]} \t D) {c[3]} +{e}+ Your ANS A/B/C/D: "
    )
    # xyz =pyip.inputStr(allowRegexes=['A','B','C','D'])
    # xyz = pyip.inputStr(f"{counter+1} To which section of the Companies Act 2013 does the below text pertain: \n    {b} \nA) {c[0]} \t B) {c[1]} \t C) {c[2]} \t D) {c[3]} \t Your ANS A/B/C/D: {e}",allowRegexes=[r'(A|B|C|D)+,r'zero''])

    xyz = xyz.capitalize()
    # print(xyz)

    if xyz == "A":
        a["selectedAnswer"] = c[0]
    elif xyz == "B":
        a["selectedAnswer"] = c[1]
    elif xyz == "C":
        a["selectedAnswer"] = c[2]
    elif xyz == "D":
        a["selectedAnswer"] = c[3]

    # print(a["selectedAnswer"])

    if a["selectedAnswer"] == a["selectedSectionNumber"]:
        score += 1


print(f"\n *****Score is {score} / 10*****")


#     q[]
#     variable1 = selectRandomSections[counter]
#     questionPaper[counter]["selectedSerial"] = variable1
#     questionPaper[counter]["selectedSectionNumber"] = section[variable1].get("sectionNumber")
#     questionPaper[counter]["selectedTextOfSection"] = section[variable1].get("textOfSection")
#     """
#     # questionPaper[counter]["selectedSelected"] = section[variable1].get("selected") #these should not even come in the dictionary!
#     # questionPaper[counter]["selectedAnsweredCorrectly"] = section[variable1].get("answeredCorrectly") #these should not even come in the dictionary!
# # #todo: the right answer will be stored as answer answer q1
# # #todo: make 3 wrong answers
# # 1 within two sections plus or minus
# # 2 within 10 sections plus or minus
# # 3 within the whole act
# # wanswer[0-2]
#     """
#     wrongAnswerSections = []
#     for counter in range(0,3):
#         wrongAnswerSections.append(selectRandomSections[counter+random.randrange(1,4)])
#         # v1 = selectRandomSections[v1]
#         # v2=selectRandomSections[counter+random.randrange(1,447)]
#         print(wrongAnswerSections[counter])
#     # questionPaper[counter]["wanswer1"] = section[variable2].get("sectionNumber")
#     # print(section[variable2].get("sectionNumber"))
#     questionPaper[counter]["wanswer1"] = selectRandomSections[wrongAnswerSections[0]]
#     print(selectRandomSections[wrongAnswerSections[0]])
#     questionPaper[counter]["wanswer2"] = selectRandomSections[wrongAnswerSections[1]]
#     questionPaper[counter]["wanswer3"] = selectRandomSections[wrongAnswerSections[2]]
#     questionPaper[counter]["selectedAnswer"] = None


# #todo: put the answer1 and wanswer[0-2] in a list as options 1
# #todo: Loop the above
#
# #todo: ask the user to provide the answer answer[0-9] =a or b or c or d
#
# #todo: provide to the user the question and answer
# print (
# the text of this section reads as
# xxx
# which section is this?
# )
# a section <wanswer 2>
# b section <answer>
# c section < wanswer 1>
# d section<w answer 3>
# your answer? Select a/b/c/d
# question 3 of 10 answered
#
# #todo: store the answer as user answerq1
#
# #todo: Last question
# <snip>
# question 10 of 10
#
# #todo: Last Question: hold your breathe for answers
# #todo: computer the answers to be correct it or wrong
#
# store all question related
# details in a dictionary
#
# #todo: if user answer = right
# answer them score q1+=1
# if user answer=any 3 of wanswer then score=0
#
# #todo: present to the user again in the same format as before, the right answer in bold, the selected answer underlined
# score for that answer right
#
# al the 10 questions should be displayed like wise
# Loop 9 times for 9 separate
# questions
